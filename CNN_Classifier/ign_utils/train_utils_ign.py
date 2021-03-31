import pandas as pd
pd.set_option('display.max_rows', 60)
import numpy as np
import math
from operator import truediv
from sklearn.utils.class_weight import compute_class_weight
from sklearn.metrics import precision_recall_fscore_support
def GetClassWeights(yTrain):
    yTrain = np.concatenate(yTrain)
    class_weights = compute_class_weight('balanced', np.unique(yTrain), yTrain)
    d_class_weights = dict(enumerate(class_weights))
    return d_class_weights

def get_confusionMatrix(ytrue, ypred, classlist):
    cols = len(classlist)
    CM = np.zeros((cols, cols)).astype(int)
    
    for ind, p_in in enumerate(ytrue):
        p_in_index = classlist.index(p_in)
        pred_val = ypred[ind]
        
        if pred_val==p_in_index:
            CM[pred_val, pred_val]+=1
        else:
            CM[p_in_index, pred_val]+=1
    
    return CM

def calc_conf_matrices(p_input, p_small, p_output, unk_list, unknown_index, classlist):
    # import pdb;pdb.set_trace()
    p_unk = [i for i in p_input if i in unk_list]
    rows = len(unk_list) 
    cols = len(classlist)
    cm1 = np.zeros((cols, cols)).astype(int)
    cm2 = np.zeros((rows, cols)).astype(int)
    o = [classlist[i] for i in p_output]
    # for ind, p_in in enumerate(p_small):
    #     p_in_index = classlist.index(p_in)
    #     pred_val = p_output[ind]
        
    #     if pred_val==p_in_index:
    #         cm1[pred_val, pred_val]+=1
    #     else:
    #         cm1[p_in_index, pred_val]+=1
    cm1 = get_confusionMatrix(p_small, p_output, classlist)
    
    for ind, p_in in enumerate(p_unk):
        p_in_index = unk_list.index(p_in)
        pred_val = p_output[ind]
        if p_in in unk_list:
            if pred_val =='unknown':
                cm2[unknown_index, unknown_index]+=1
            else:
                cm2[p_in_index, pred_val]+=1
        else:
            cm2[p_in_index, pred_val]+=1
    df1, df2, metrics_df, avgmetrics, df_norm, norm_metrics_df, norm_average_metrics = Get_dataframes(cm1, cm2, p_small, p_output, classlist, unk_list)
    
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.set_option('display.max_colwidth', -1)
    print('\n***CM1***:')
    print(df1)
    print(metrics_df)
    print (avgmetrics.to_string(header=False))

    print('\n***CM1 Normalized***:')
    print(df_norm)
    
    print('\n***Accuracy Metrics CM1 Norm***:')
    print(norm_metrics_df)
    
    print('\n***Average Accuracy Metrics CM1 Norm***:')
    print (norm_average_metrics.to_string(header=False))

    print('\n***CM2***:')
    print(df2)
    
    return df1, df2, metrics_df, avgmetrics, df_norm, norm_metrics_df, norm_average_metrics

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       startcol=None, truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # if startcol is None and sheet_name in writer.book.sheetnames:
        #     startcol = writer.book[sheet_name].max_col
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()

def export_to_excel(df1, df2, metrics_df, avgmetrics, df_norm, norm_metrics_df, norm_average_metrics, filename, model_name):
    startcol = 0
    start_row = 0
    title = pd.DataFrame(['Confusion Matrix1 RAW'])
    append_df_to_excel(filename, title, header=False, sheet_name=model_name)
    start_row+=2
    append_df_to_excel(filename, df1, sheet_name=model_name, startrow=start_row)
    start_row+=len(df1) + 3
    # startcol += (len(df1.columns)+4)
   
   
    # title = pd.DataFrame(['Accuracy Metrics'])
    # append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, startcol=startcol)
    
    append_df_to_excel(filename, metrics_df, sheet_name=model_name, startrow=start_row, startcol=1)
    # startcol+=(len(metrics_df.columns)+4)
    start_row+=len(metrics_df)+3
    # title = pd.DataFrame(['Average Accuracy Metrics'])
    # append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, startcol=startcol)
    
    append_df_to_excel(filename, avgmetrics, sheet_name=model_name, startrow=start_row, header=False, startcol=1)
    startcol+=(len(avgmetrics.columns)+4)
    start_row+=(len(avgmetrics)+3)
    startcol = 0

    
    title = pd.DataFrame(['Confusion Matrix1 normalized'])
    append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, header=False, startcol=startcol)
    start_row+=2
    append_df_to_excel(filename, df_norm, sheet_name=model_name, startrow=start_row, startcol=startcol)
    startcol+=(len(df_norm.columns)+4)
    startcol = 0
    start_row+=len(df_norm)+3
    # title = pd.DataFrame(['Accuracy Metrics(norm)'])
    # append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, startcol=startcol)
    
    append_df_to_excel(filename, norm_metrics_df, sheet_name=model_name, startrow=start_row, startcol=1)
    # startcol+=(len(norm_metrics_df.columns)+4)
    start_row+=(len(norm_metrics_df)+3)

    # title = pd.DataFrame(['Average Accuracy Metrics(norm)'])
    # append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, startcol=startcol)
    
    append_df_to_excel(filename, norm_average_metrics, sheet_name=model_name, startrow=start_row, header=False, startcol=1)
    start_row+=(len(df_norm)+2)
    startcol = 0
    
    title = pd.DataFrame(['Confusion Matrix2'])
    append_df_to_excel(filename, title, sheet_name=model_name, startrow=start_row, header=False, startcol=startcol)
    start_row+=2
    append_df_to_excel(filename, df2, sheet_name=model_name, startrow=start_row, startcol=startcol)


def export_to_excelOld(df1, df2, metrics_df, avgmetrics, df_norm, norm_metrics_df, norm_average_metrics, filename, model_name):
    title = pd.DataFrame(['Confusion Matrix RAW'])
    start_row = 2
    append_df_to_excel(filename, df1, sheet_name=model_name)
    start_row += len(df1)+5
    append_df_to_excel(filename, metrics_df, sheet_name=model_name, startrow=start_row)
    start_row+=len(metrics_df)+5
    append_df_to_excel(filename, avgmetrics, sheet_name=model_name, startrow=start_row, header=False)
    start_row+=len(avgmetrics)+5
    append_df_to_excel(filename, df2, sheet_name=model_name, startrow=start_row)


def get_normalised_df(df, columns, scale_val=None):
    df["sum"] = df.sum(axis=1)
    df.loc[:,columns] = df.loc[:,columns].div(df["sum"], axis=0)
    df.insert(loc=0, column='tp', value=columns)
    if scale_val is not None:
        df.loc[:,columns] = df.loc[:,columns].mul(scale_val, axis=0)
    return df[columns].round()

def get_precision(true_pos, false_pos):
    prec = (true_pos / (true_pos + false_pos))
    prec = [np.round(i, 2) for i in prec]
    prec = [0 if math.isnan(x) is True else x for x in prec]
    return np.array(prec)

def get_recall(true_pos, false_neg):
    rec = (true_pos / (true_pos + false_neg))
    rec = [np.round(i, 2) for i in rec]
    return np.array(rec)

def get_f1_score(prec, rec):
    f1_score = [calc_f1Score(pr, re) for pr, re in zip(prec, rec)]
    f1_score = [np.round(i, 2) for i in f1_score]
    f1_score = [0 if math.isnan(x) is True else x for x in f1_score[0]]
    return np.array(f1_score)

def get_precision_recall_from_df(df, columns):
    cm = df[columns].to_numpy()
    width = cm.shape[0]
    true_pos = np.diag(cm)
    false_pos = np.sum(cm, axis=0) - true_pos
    false_neg = np.sum(cm, axis=1) - true_pos
    
    prec = get_precision(true_pos, false_pos).reshape(1, width)
    rec = get_recall(true_pos, false_neg).reshape(1, width)
    f1_score = get_f1_score(prec, rec).reshape(1, width)

    precision_list = ['Precision', 'Recall', 'F1Score']
    norm_metrics_df = pd.DataFrame([list(prec[0]), list(rec[0]), list(f1_score[0])], precision_list, columns=columns) 
    mean_vals = norm_metrics_df.mean(axis=1)
    norm_average_metrics = pd.DataFrame([mean_vals[0], mean_vals[1], mean_vals[2]], ['Average Precision(norm)', 'Average Recall(norm)', 'Average F1Score(norm)'])
    return norm_metrics_df, norm_average_metrics

def Get_dataframes(cm1, cm2, ytrue, ypred, labels, classes):
    """pretty print for confusion matrixes"""
    # import pdb;pdb.set_trace()
    width = cm1.shape[0]
    true_pos = np.diag(cm1)
    false_pos = np.sum(cm1, axis=0) - true_pos
    false_neg = np.sum(cm1, axis=1) - true_pos

    prec = get_precision(true_pos, false_pos).reshape(1, width)
    rec = get_recall(true_pos, false_neg).reshape(1, width)
    f1_score = get_f1_score(prec, rec).reshape(1, width)
    precision_list = ['Precision', 'Recall', 'F1Score']
    
    # import pdb;pdb.set_trace()
    metrics_df = pd.DataFrame([list(prec[0]), list(rec[0]), list(f1_score[0])], precision_list, columns=labels) 
    mean_vals = metrics_df.mean(axis=1)
    
    df1 = pd.DataFrame(cm1, columns=labels)
    df_norm = get_normalised_df(df1.copy(), labels, scale_val=100)
    
    norm_metrics_df, norm_average_metrics = get_precision_recall_from_df(df_norm, labels)
    # import pdb;pdb.set_trace()
    df1.insert(loc=0, column='tp', value=labels)
    df_norm.insert(loc=0, column='tp', value=labels)
    metrics_df.loc['support'] = np.array(df1.sum(axis=1))
    norm_metrics_df.loc['support'] = np.array(df1.sum(axis=1))
    df2 = pd.DataFrame(cm2, columns=labels)
    df2.insert(loc=0, column='tp', value=classes)
    
    all_column_list = df2.columns.tolist()
    remove_list = [all_column_list[0], all_column_list[-1]]
    sort_by_list = [i for i in all_column_list if i not in remove_list]
    
    df2['false_sum'] = df2[sort_by_list].sum(axis=1)

    sorted_df2 = df2.sort_values(['false_sum'], ascending=False)
    sorted_df2 = sorted_df2[all_column_list].reset_index(drop=True)
    average_metrics = pd.DataFrame([mean_vals[0], mean_vals[1], mean_vals[2]], ['Average Precision', 'Average Recall', 'Average F1Score'])
    return df1, sorted_df2, metrics_df, average_metrics, df_norm, norm_metrics_df, norm_average_metrics

def calc_f1Score(pr, re):
    return 2* ((re * pr)/ (re + pr))
